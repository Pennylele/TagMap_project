#! /usr/bin/env python3
# TagMap.py is used for disect the filter xml file exported from memoQ and put
# those needed content into an excel.

import os, openpyxl
from bs4 import BeautifulSoup
from openpyxl.styles import Font, Color, PatternFill, NamedStyle
from openpyxl.styles import colors
from openpyxl import Workbook
import io

cwd = os.getcwd()


def func(outputFile):
    wbOut = openpyxl.Workbook()
    sheetOut = wbOut.active
    
    r = 1
    for file in os.listdir(cwd):
        if file.endswith('.mqres'):
            print('The filter file is ' + file)
            f = io.open(file, encoding = 'utf-8', mode = "r+")
            soup = BeautifulSoup(f, 'lxml')

            #find all the nontranslatable tags
            
            TagsNO = soup.find_all(attrs={"nontranslated": "no"})
            for Tag in TagsNO:
                names = Tag.get('name')
                f = open('output.txt', 'a')
                f.write(str(names) + "\n")
            f = open('output.txt', 'r+')
            lines = f.readlines()
            lines.sort(key=str.lower)
            for line in lines:
                sheetOut.cell(row=1, column=1).value = "Translatable tags"
                #sheetOut['A1'].fill = blackFill
                sheetOut.cell(row=r+1, column=1).value = line
                r += 1

            #find all the translatable tags
            r = 1
            TagsYES = soup.find_all(attrs={"nontranslated": "yes"})
            for Tag in TagsYES:
                names = Tag.get('name')
                f = open('output2.txt', 'a')
                f.write(str(names) + "\n")
            f = open('output2.txt', 'r+')
            lines = f.readlines()
            lines.sort(key=str.lower)
            for line in lines:
                sheetOut.cell(row=1, column=2).value = "Non-translatable tags"
                sheetOut.cell(row=r+1, column=2).value = line
                r += 1

            #find translatable attributes
            r = 3
            AttrYES = soup.findAll(attrs={'translatable': 'yes'})
            result = dict()
            i, j = 3, 1
            for items in AttrYES:
                result[items.get('name')] = result.get(items.get('name'),[])+[items.parent.parent.get('name')]
            for keys,values in result.items():
                sheetOut.cell(column=i, row=1, value='Translatable attribute: '+ keys)
                for row in range(len(values)):
                    values.sort(key=str.lower)
                    sheetOut.cell(column=i, row=j+1, value=values[row])
                    j+=1
                j=1
                i+=1

    highlight = NamedStyle(name="highlight")
    highlight.font = Font(bold=True, name="Arial")

    for cell in sheetOut["1:1"]:
        cell.style = highlight

    sheetOut.column_dimensions['A'].width = 30
    sheetOut.column_dimensions['B'].width = 30
    sheetOut.column_dimensions['C'].width = 35
    sheetOut.column_dimensions['D'].width = 35
    sheetOut.column_dimensions['E'].width = 35
    sheetOut.column_dimensions['F'].width = 35
    sheetOut.column_dimensions['G'].width = 35
    sheetOut.column_dimensions['H'].width = 35
    sheetOut.column_dimensions['I'].width = 35
    sheetOut.column_dimensions['J'].width = 35
    sheetOut.column_dimensions['K'].width = 35
    sheetOut.column_dimensions['L'].width = 35
    sheetOut.column_dimensions['M'].width = 35
    sheetOut.column_dimensions['N'].width = 35
    sheetOut.column_dimensions['O'].width = 35
    sheetOut.column_dimensions['P'].width = 35
    sheetOut.column_dimensions['Q'].width = 35
    sheetOut.column_dimensions['R'].width = 35
    sheetOut.column_dimensions['S'].width = 35
    sheetOut.column_dimensions['T'].width = 35
    sheetOut.column_dimensions['U'].width = 35
    sheetOut.column_dimensions['V'].width = 35
    sheetOut.column_dimensions['W'].width = 35
    sheetOut.column_dimensions['X'].width = 35
    sheetOut.column_dimensions['Y'].width = 35
    sheetOut.column_dimensions['Z'].width = 35
                            
    
    wbOut.save('XML_Tagmap.xlsx')

    for file in os.listdir(cwd):
        if file.endswith('.txt'):
            os.remove(file)

outputFile = os.listdir(cwd)
func(outputFile)
